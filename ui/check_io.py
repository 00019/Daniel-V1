import csv as _csv
import shutil
import sqlite3
from datetime import datetime

from PySide6.QtWidgets import QFileDialog, QMessageBox

from database.table import Table


class CheckIOMixin:

    def _get_all_checks_data(self):
        self.table.load_from_db()
        all_checks = list(Table.checks)
        data = []
        for c in all_checks:
            data.append({
                "ID":        c.get("id"),
                "Check #":   c.get("check_number"),
                "Cleared":   self._to_bool(c.get("cleared")),
                "Payee":     self._to_str(c.get("payee")),
                "Drawer":    self._to_str(c.get("drawer")),
                "Drawee":    self._to_str(c.get("drawee")),
                "Amount":    self._to_float(c.get("amount")),
                "Date":      self._format_date_text(c.get("date")),
                "Account":   self._to_str(c.get("account")),
                "Routing #": self._to_str(c.get("routing")),
                "Memo":      self._to_str(c.get("memo")),
            })
        return data

    def _get_current_table_data(self):
        model = self.table_view.model()
        if not model:
            return []
        data = []
        for row in range(model.rowCount()):
            row_dict = {}
            for col, field in enumerate(self.COLUMNS):
                if field == "Cleared":
                    item = model.item(row, self.FIELD_TO_COL["cleared"])
                    from PySide6.QtCore import Qt
                    row_dict[field] = (item.checkState() == Qt.Checked) if item else False
                else:
                    idx = model.index(row, col)
                    value = idx.data()
                    if field == "Amount" and isinstance(value, str):
                        value = value.replace("$", "")
                    row_dict[field] = value
            data.append(row_dict)
        return data

    def _get_data_from_ids(self, ids: list) -> list:
        self.table.load_from_db()
        id_set = set(ids)
        data = []
        for c in Table.checks:
            if c.get("id") in id_set:
                data.append({
                    "ID":        c.get("id"),
                    "Check #":   c.get("check_number"),
                    "Cleared":   self._to_bool(c.get("cleared")),
                    "Payee":     self._to_str(c.get("payee")),
                    "Drawer":    self._to_str(c.get("drawer")),
                    "Drawee":    self._to_str(c.get("drawee")),
                    "Amount":    self._to_float(c.get("amount")),
                    "Date":      self._format_date_text(c.get("date")),
                    "Account":   self._to_str(c.get("account")),
                    "Routing #": self._to_str(c.get("routing")),
                    "Memo":      self._to_str(c.get("memo")),
                })
        return data

    def _export_all_to(self, format_type: str):
        data = self._get_all_checks_data()
        self._perform_export(data, format_type)

    def _export_displayed_to(self, format_type: str):
        data = self._get_current_table_data()
        self._perform_export(data, format_type)

    def _export_selected_to(self, format_type: str):
        ids = self._get_selected_ids()
        if not ids:
            QMessageBox.warning(self, "No Selection", "Please select at least one check to export.")
            return
        data = self._get_data_from_ids(ids)
        self._perform_export(data, format_type)

    def _perform_export(self, data: list, format_type: str):
        if not data:
            QMessageBox.information(self, "Nothing to Export", "No checks to export.")
            return
        if format_type == "csv":
            self._export_to_csv(data)
        elif format_type == "excel":
            self._export_to_excel(data)
        elif format_type == "qif":
            self._export_to_qif(data)
        elif format_type == "db":
            self._export_to_db(data)
        else:
            raise ValueError(f"Unknown format: {format_type}")

    def _export_to_csv(self, data: list):
        save_path, _ = QFileDialog.getSaveFileName(self, "Export to CSV", "", "CSV Files (*.csv)")
        if not save_path:
            return
        if not save_path.endswith(".csv"):
            save_path += ".csv"
        fieldnames = ["ID", "Check #", "Cleared", "Payee", "Drawer", "Drawee", "Amount", "Date", "Account", "Routing #", "Memo"]
        try:
            with open(save_path, "w", newline="", encoding="utf-8") as fh:
                writer = _csv.DictWriter(fh, fieldnames=fieldnames)
                writer.writeheader()
                for row in data:
                    row_copy = row.copy()
                    row_copy["Cleared"] = "Yes" if row_copy.get("Cleared") else "No"
                    writer.writerow(row_copy)
            QMessageBox.information(self, "Export Complete", f"Exported {len(data)} rows to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_to_excel(self, data: list):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.critical(self, "Missing Library", "Install openpyxl: pip install openpyxl")
            return
        if not data:
            QMessageBox.information(self, "Nothing to Export", "No checks to export.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Cheques"
        headers = ["Check #", "Cleared", "Payee", "Drawer", "Drawee", "Amount", "Date", "Account", "Routing #", "Memo"]
        ws.append(headers)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_dict in data:
            ws.append([
                row_dict.get("Check #", ""),
                "Yes" if row_dict.get("Cleared") else "No",
                row_dict.get("Payee", ""),
                row_dict.get("Drawer", ""),
                row_dict.get("Drawee", ""),
                row_dict.get("Amount", ""),
                row_dict.get("Date", ""),
                row_dict.get("Account", ""),
                row_dict.get("Routing #", ""),
                row_dict.get("Memo", ""),
            ])
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if not save_path:
            return
        if not save_path.endswith(".xlsx"):
            save_path += ".xlsx"
        try:
            wb.save(save_path)
            QMessageBox.information(self, "Export Complete", f"Exported {len(data)} rows to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to save Excel file:\n{e}")

    def _export_to_qif(self, data: list):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save QIF", "", "QIF Files (*.qif)")
        if not save_path:
            return
        if not save_path.endswith(".qif"):
            save_path += ".qif"
        lines = ["!Type:Bank"]
        for row in data:
            date_str = row.get("Date", "")
            amount = row.get("Amount", 0)
            payee = row.get("Payee", "")
            memo = row.get("Memo", "")
            check_num = row.get("Check #", "")
            try:
                amt_val = float(amount) if amount else 0.0
            except Exception:
                amt_val = 0.0
            lines.append(f"D{date_str}")
            if check_num:
                lines.append(f"N{check_num}")
            lines.append(f"P{payee}")
            lines.append(f"T{-amt_val:.2f}")
            if memo:
                lines.append(f"M{memo}")
            lines.append("^")
        try:
            with open(save_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            QMessageBox.information(self, "Export Complete", f"Exported {len(data)} checks to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_to_db(self, data: list):
        all_checks = self._get_all_checks_data()
        if len(data) == len(all_checks):
            save_path, _ = QFileDialog.getSaveFileName(self, "Save Database Copy", "", "SQLite Database (*.db)")
            if not save_path:
                return
            if not save_path.endswith(".db"):
                save_path += ".db"
            try:
                shutil.copy2(self.db_path, save_path)
                QMessageBox.information(self, "Export Complete", f"Database copied to:\n{save_path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", str(e))
            return
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Filtered Database", "", "SQLite Database (*.db)")
        if not save_path:
            return
        if not save_path.endswith(".db"):
            save_path += ".db"
        try:
            conn = sqlite3.connect(save_path)
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS checks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    check_number INTEGER NOT NULL,
                    cleared INTEGER NOT NULL,
                    payee TEXT NOT NULL,
                    drawer TEXT,
                    drawee TEXT,
                    amount REAL NOT NULL,
                    date TEXT NOT NULL,
                    account TEXT,
                    routing TEXT,
                    memo TEXT
                )
            ''')
            for row in data:
                cleared_val = 1 if row.get("Cleared") else 0
                cursor.execute('''
                    INSERT INTO checks
                    (check_number, cleared, payee, drawer, drawee, amount, date, account, routing, memo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    row.get("Check #"),
                    cleared_val,
                    row.get("Payee"),
                    row.get("Drawer"),
                    row.get("Drawee"),
                    row.get("Amount"),
                    row.get("Date"),
                    row.get("Account"),
                    row.get("Routing #"),
                    row.get("Memo"),
                ))
            conn.commit()
            conn.close()
            QMessageBox.information(self, "Export Complete", f"Filtered database created at:\n{save_path}\n({len(data)} checks)")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _import_from_csv(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select CSV File", "", "CSV Files (*.csv)")
        if not file_path:
            return
        imported = 0
        errors = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = _csv.DictReader(f)
                for row_num, row in enumerate(reader, start=2):
                    try:
                        check_number = int(row.get("Check #", ""))
                        cleared = row.get("Cleared", "").lower() == "yes"
                        payee = row.get("Payee", "").strip()
                        drawer = row.get("Drawer", "").strip()
                        drawee = row.get("Drawee", "").strip()
                        amount = float(row.get("Amount", 0))
                        date_str = row.get("Date", "")
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                        account = row.get("Account", "").strip()
                        routing = row.get("Routing #", "").strip()
                        memo = row.get("Memo", "").strip()
                        data = {
                            "check_number": check_number,
                            "cleared": cleared,
                            "payee": payee,
                            "drawer": drawer,
                            "drawee": drawee,
                            "amount": amount,
                            "date": date_obj,
                            "account": account,
                            "routing": routing,
                            "memo": memo,
                        }
                        self.table.add_check(data)
                        imported += 1
                    except Exception as e:
                        errors.append(f"Row {row_num}: {e}")
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to read CSV:\n{e}")
            return
        self.refresh_all()
        msg = f"Imported {imported} checks."
        if errors:
            msg += f"\n\n{len(errors)} error(s):\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                msg += f"\n... and {len(errors)-5} more."
        QMessageBox.information(self, "Import Complete", msg)

    def _import_from_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.critical(self, "Missing Library", "Install openpyxl: pip install openpyxl")
            return
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value:
                    headers[cell.value.strip()] = col_idx
            required = ["Check #", "Payee", "Amount", "Date"]
            for r in required:
                if r not in headers:
                    raise ValueError(f"Missing required column '{r}' in Excel file")
            imported = 0
            errors = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if all(cell is None for cell in row):
                        continue
                    check_number = int(row[headers["Check #"]-1]) if row[headers["Check #"]-1] is not None else 0
                    cleared = str(row[headers["Cleared"]-1]).lower() == "yes" if headers.get("Cleared") and row[headers["Cleared"]-1] else False
                    payee = str(row[headers["Payee"]-1]).strip() if row[headers["Payee"]-1] else ""
                    drawer = str(row[headers["Drawer"]-1]).strip() if headers.get("Drawer") and row[headers["Drawer"]-1] else ""
                    drawee = str(row[headers["Drawee"]-1]).strip() if headers.get("Drawee") and row[headers["Drawee"]-1] else ""
                    amount = float(row[headers["Amount"]-1]) if row[headers["Amount"]-1] else 0.0
                    date_str = str(row[headers["Date"]-1]).strip() if row[headers["Date"]-1] else ""
                    account = str(row[headers["Account"]-1]).strip() if headers.get("Account") and row[headers["Account"]-1] else ""
                    routing = str(row[headers["Routing #"]-1]).strip() if headers.get("Routing #") and row[headers["Routing #"]-1] else ""
                    memo = str(row[headers["Memo"]-1]).strip() if headers.get("Memo") and row[headers["Memo"]-1] else ""
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                    data = {
                        "check_number": check_number,
                        "cleared": cleared,
                        "payee": payee,
                        "drawer": drawer,
                        "drawee": drawee,
                        "amount": amount,
                        "date": date_obj,
                        "account": account,
                        "routing": routing,
                        "memo": memo,
                    }
                    self.table.add_check(data)
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {e}")
            self.refresh_all()
            msg = f"Imported {imported} checks from Excel."
            if errors:
                msg += f"\n\n{len(errors)} error(s):\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n... and {len(errors)-5} more."
            QMessageBox.information(self, "Import Complete", msg)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to read Excel file:\n{e}")

    def _import_from_qif(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select QIF File", "", "QIF Files (*.qif)")
        if not file_path:
            return
        imported = 0
        errors = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            current = {}
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if line == "!Type:Bank":
                    continue
                if line == "^":
                    if current:
                        try:
                            check_number = current.get("N", "")
                            check_number = int(check_number) if check_number else 0
                            payee = current.get("P", "")
                            amount = abs(float(current.get("T", "0")))
                            date_str = current.get("D", "")
                            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                            memo = current.get("M", "")
                            data = {
                                "check_number": check_number,
                                "cleared": False,
                                "payee": payee,
                                "drawer": "",
                                "drawee": "",
                                "routing": "",
                                "amount": amount,
                                "date": date_obj,
                                "account": "",
                                "memo": memo,
                            }
                            self.table.add_check(data)
                            imported += 1
                        except Exception as e:
                            errors.append(f"Transaction error: {e}")
                    current = {}
                    continue
                if line[0] in "DNPTM":
                    current[line[0]] = line[1:]
            self.refresh_all()
            msg = f"Imported {imported} checks."
            if errors:
                msg += f"\n\n{len(errors)} error(s):\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n... and {len(errors)-5} more."
            QMessageBox.information(self, "Import Complete", msg)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to read QIF:\n{e}")

    def _import_from_db(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select SQLite Database", "", "Database Files (*.db)")
        if not file_path:
            return
        try:
            src_conn = sqlite3.connect(file_path)
            src_cursor = src_conn.cursor()
            src_cursor.execute("SELECT check_number, cleared, payee, drawer, drawee, amount, date, account, routing, memo FROM checks")
            rows = src_cursor.fetchall()
            src_conn.close()
            imported = 0
            errors = []
            for row in rows:
                try:
                    check_number, cleared, payee, drawer, drawee, amount, date_str, account, routing, memo = row
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                    data = {
                        "check_number": check_number,
                        "cleared": bool(cleared),
                        "payee": payee,
                        "drawer": drawer or "",
                        "drawee": drawee or "",
                        "amount": amount,
                        "date": date_obj,
                        "account": account or "",
                        "routing": routing or "",
                        "memo": memo or "",
                    }
                    self.table.add_check(data)
                    imported += 1
                except Exception as e:
                    errors.append(str(e))
            self.refresh_all()
            msg = f"Imported {imported} checks from database."
            if errors:
                msg += f"\n\n{len(errors)} error(s):\n" + "\n".join(errors[:5])
            QMessageBox.information(self, "Import Complete", msg)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to read database:\n{e}")

    def _append_to_db(self, data: list, description: str):
        if not data:
            QMessageBox.information(self, "Nothing to Append", f"No checks to append ({description}).")
            return
        target_path, _ = QFileDialog.getOpenFileName(self, "Select Target Database", "", "SQLite Database (*.db)")
        if not target_path:
            return
        try:
            conn = sqlite3.connect(target_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='checks'")
            if not cursor.fetchone():
                conn.close()
                raise ValueError("Target database does not have a 'checks' table.")
            cursor.execute("PRAGMA table_info(checks)")
            columns = {col[1] for col in cursor.fetchall()}
            required = {"check_number", "cleared", "payee", "drawer", "drawee", "amount", "date", "account", "routing", "memo"}
            if not required.issubset(columns):
                missing = required - columns
                conn.close()
                raise ValueError(f"Target 'checks' table is missing columns: {missing}")
            inserted = 0
            for row in data:
                cursor.execute('''
                    INSERT INTO checks
                    (check_number, cleared, payee, drawer, drawee, amount, date, account, routing, memo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    row.get("Check #"),
                    1 if row.get("Cleared") else 0,
                    row.get("Payee", ""),
                    row.get("Drawer", ""),
                    row.get("Drawee", ""),
                    row.get("Amount", 0.0),
                    row.get("Date", ""),
                    row.get("Account", ""),
                    row.get("Routing #", ""),
                    row.get("Memo", ""),
                ))
                inserted += 1
            conn.commit()
            conn.close()
            QMessageBox.information(self, "Append Complete", f"Appended {inserted} checks to:\n{target_path}")
        except Exception as e:
            QMessageBox.critical(self, "Append Error", str(e))

    def _append_all_to_db(self):
        data = self._get_all_checks_data()
        self._append_to_db(data, "all")

    def _append_displayed_to_db(self):
        data = self._get_current_table_data()
        self._append_to_db(data, "displayed")

    def _append_selected_to_db(self):
        ids = self._get_selected_ids()
        if not ids:
            QMessageBox.warning(self, "No Selection", "Please select at least one check to append.")
            return
        data = self._get_data_from_ids(ids)
        self._append_to_db(data, "selected")
